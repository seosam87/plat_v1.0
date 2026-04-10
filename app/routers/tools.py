"""Tools router: TOOL_REGISTRY dispatch, 7 route handlers for Phase 24-25 tools."""
from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime, timezone

from celery import chain as celery_chain
from fastapi import APIRouter, Depends, Form, HTTPException, Request, Response, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl

from app.dependencies import get_db
from app.auth.dependencies import get_current_user
from app.models.user import User
from app.rate_limit import limiter
from app.template_engine import templates

router = APIRouter(prefix="/ui/tools", tags=["tools"])

# Tool metadata — models and tasks populated by Plans 02-04.
# form_field: HTML textarea name attribute
# input_col: Job model column for storing input list
# count_col: Job model column for input count
TOOL_REGISTRY: dict[str, dict] = {
    "commercialization": {
        "name": "Проверка коммерциализации",
        "description": "Анализ коммерциализации поисковой выдачи по ключевым фразам",
        "input_type": "phrases",
        "form_field": "phrases",
        "input_col": "input_phrases",
        "count_col": "phrase_count",
        "limit": 200,
        "cta": "Проверить коммерциализацию",
        "slug": "commercialization",
        "has_domain_field": False,
    },
    "meta-parser": {
        "name": "Парсер мета-тегов",
        "description": "Извлечение title, description, H1, H2 и canonical с указанных URL",
        "input_type": "urls",
        "form_field": "urls",
        "input_col": "input_urls",
        "count_col": "url_count",
        "limit": 500,
        "cta": "Запустить парсинг",
        "slug": "meta-parser",
        "has_domain_field": False,
    },
    "relevant-url": {
        "name": "Поиск релевантного URL",
        "description": "Поиск релевантных страниц домена в ТОП-10 Яндекса по ключевым фразам",
        "input_type": "phrases",
        "form_field": "phrases",
        "input_col": "input_phrases",
        "count_col": "phrase_count",
        "limit": 100,
        "cta": "Найти релевантные URL",
        "slug": "relevant-url",
        "has_domain_field": True,
    },
    "brief": {
        "name": "Копирайтерское ТЗ",
        "description": "Анализ ТОП-10 выдачи: сбор H2, тематических слов и статистики страниц",
        "input_type": "phrases",
        "form_field": "phrases",
        "input_col": "input_phrases",
        "count_col": "phrase_count",
        "limit": 30,
        "cta": "Составить ТЗ",
        "slug": "brief",
        "has_domain_field": False,
        "has_region_field": True,
        "export_only_xlsx": True,
    },
    "wordstat-batch": {
        "name": "Частотность (пакет)",
        "description": "Пакетная проверка частотности по Яндекс.Wordstat (до 1000 фраз)",
        "input_type": "phrases",
        "form_field": "phrases",
        "input_col": "input_phrases",
        "count_col": "phrase_count",
        "limit": 1000,
        "cta": "Проверить частотность",
        "slug": "wordstat-batch",
        "has_domain_field": False,
        "export_only_xlsx": True,
        "needs_oauth": "wordstat",
    },
    "paa": {
        "name": "PAA-парсер",
        "description": "Извлечение вопросов из блоков «Частые вопросы» и «Похожие запросы» Яндекса",
        "input_type": "phrases",
        "form_field": "phrases",
        "input_col": "input_phrases",
        "count_col": "phrase_count",
        "limit": 50,
        "cta": "Получить вопросы",
        "slug": "paa",
        "has_domain_field": False,
    },
}

# Export column headers per tool
_EXPORT_HEADERS: dict[str, list[str]] = {
    "commercialization": ["Фраза", "Коммерциализация", "Интент", "Геозависимость", "Локализация"],
    "meta-parser": ["URL", "Статус", "Title", "Description", "H1", "Robots", "Canonical"],
    "relevant-url": ["Фраза", "URL", "Позиция", "Топ-3 конкурента"],
    "brief": ["Показатель", "Значение"],  # Brief uses sectioned XLSX export
    "wordstat-batch": ["Фраза", "Точная частота", "Широкая частота"],  # Monthly in Sheet 2
    "paa": ["Фраза", "Вопрос", "Блок"],
}


def _get_tool_models(slug: str):
    """Lazy-import tool models to avoid circular imports."""
    if slug == "commercialization":
        from app.models.commerce_check_job import CommerceCheckJob, CommerceCheckResult
        return CommerceCheckJob, CommerceCheckResult
    elif slug == "meta-parser":
        from app.models.meta_parse_job import MetaParseJob, MetaParseResult
        return MetaParseJob, MetaParseResult
    elif slug == "relevant-url":
        from app.models.relevant_url_job import RelevantUrlJob, RelevantUrlResult
        return RelevantUrlJob, RelevantUrlResult
    elif slug == "brief":
        from app.models.brief_job import BriefJob, BriefResult
        return BriefJob, BriefResult
    elif slug == "wordstat-batch":
        from app.models.wordstat_batch_job import WordstatBatchJob, WordstatBatchResult
        return WordstatBatchJob, WordstatBatchResult
    elif slug == "paa":
        from app.models.paa_job import PAAJob, PAAResult
        return PAAJob, PAAResult
    raise HTTPException(status_code=404, detail="Unknown tool")


def _get_tool_task(slug: str):
    """Lazy-import tool task to avoid circular imports."""
    if slug == "commercialization":
        from app.tasks.commerce_check_tasks import run_commerce_check
        return run_commerce_check
    elif slug == "meta-parser":
        from app.tasks.meta_parse_tasks import run_meta_parse
        return run_meta_parse
    elif slug == "relevant-url":
        from app.tasks.relevant_url_tasks import run_relevant_url
        return run_relevant_url
    elif slug == "brief":
        # Brief uses a 4-step chain — caller must handle this specially
        # Returns None to signal chain dispatch in tool_submit
        return None
    elif slug == "wordstat-batch":
        from app.tasks.wordstat_batch_tasks import run_wordstat_batch
        return run_wordstat_batch
    elif slug == "paa":
        from app.tasks.paa_tasks import run_paa
        return run_paa
    raise HTTPException(status_code=404, detail="Unknown tool")


def _result_to_row(result, slug: str) -> list:
    """Convert a result ORM row to a list of cell values for export.

    Brief results use a sectioned XLSX export and should never reach this path.
    """
    if slug == "brief":
        # Brief export uses a multi-section XLSX — this path should not be reached
        return []
    if slug == "commercialization":
        return [
            getattr(result, "phrase", ""),
            getattr(result, "commercialization", ""),
            getattr(result, "intent", ""),
            "Да" if getattr(result, "geo_dependent", False) else "Нет",
            "Да" if getattr(result, "localized", False) else "Нет",
        ]
    elif slug == "meta-parser":
        return [
            getattr(result, "input_url", ""),
            getattr(result, "status_code", ""),
            getattr(result, "title", ""),
            getattr(result, "meta_description", ""),
            getattr(result, "h1", ""),
            getattr(result, "robots", ""),
            getattr(result, "canonical", ""),
        ]
    elif slug == "relevant-url":
        competitors = getattr(result, "top_competitors", None) or []
        return [
            getattr(result, "phrase", ""),
            getattr(result, "url", "") or "Не найден",
            getattr(result, "position", "") or "",
            ", ".join(competitors) if competitors else "",
        ]
    elif slug == "wordstat-batch":
        return [
            getattr(result, "phrase", ""),
            getattr(result, "freq_exact", ""),
            getattr(result, "freq_broad", ""),
        ]
    elif slug == "paa":
        return [
            getattr(result, "phrase", ""),
            getattr(result, "question", ""),
            getattr(result, "source_block", ""),
        ]
    return []


def _check_oauth_token_sync(needs_oauth: str) -> str | None:
    """Synchronous helper: load OAuth token for the given service type.

    Used by tool_landing to determine whether to show the OAuth warning banner.
    Runs in a thread executor to avoid blocking the async event loop.
    """
    from app.database import get_sync_db
    from app.services.batch_wordstat_service import check_wordstat_oauth_token

    if needs_oauth == "wordstat":
        with get_sync_db() as db:
            return check_wordstat_oauth_token(db)
    return None


# ---------------------------------------------------------------------------
# 1. GET / — tools index: card grid with job count badges
# ---------------------------------------------------------------------------
@router.get("/", response_class=HTMLResponse, name="tools_index")
async def tools_index(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> HTMLResponse:
    """Tools index page with tool cards and per-user job count badges."""
    from app.models.brief_job import BriefJob
    from app.models.commerce_check_job import CommerceCheckJob
    from app.models.meta_parse_job import MetaParseJob
    from app.models.paa_job import PAAJob
    from app.models.relevant_url_job import RelevantUrlJob
    from app.models.wordstat_batch_job import WordstatBatchJob

    job_counts: dict[str, int] = {}
    for slug, model_cls in [
        ("commercialization", CommerceCheckJob),
        ("meta-parser", MetaParseJob),
        ("relevant-url", RelevantUrlJob),
        ("brief", BriefJob),
        ("wordstat-batch", WordstatBatchJob),
        ("paa", PAAJob),
    ]:
        result = await db.execute(
            select(func.count(model_cls.id)).where(model_cls.user_id == user.id)
        )
        job_counts[slug] = result.scalar() or 0

    tools = []
    for slug, info in TOOL_REGISTRY.items():
        tools.append({**info, "job_count": job_counts.get(slug, 0)})

    return templates.TemplateResponse(
        request,
        "tools/index.html",
        {"tools": tools},
    )


# ---------------------------------------------------------------------------
# 2. GET /{slug}/ — tool landing: input form + previous jobs list
# ---------------------------------------------------------------------------
@router.get("/{slug}/", response_class=HTMLResponse, name="tool_landing")
async def tool_landing(
    slug: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> HTMLResponse:
    if slug not in TOOL_REGISTRY:
        raise HTTPException(status_code=404, detail="Unknown tool")
    registry = TOOL_REGISTRY[slug]
    JobModel, _ResultModel = _get_tool_models(slug)

    stmt = (
        select(JobModel)
        .where(JobModel.user_id == user.id)
        .order_by(JobModel.created_at.desc())
        .limit(20)
    )
    result = await db.execute(stmt)
    jobs = result.scalars().all()

    # OAuth warning: check if required token is configured
    oauth_warning = False
    if registry.get("needs_oauth"):
        from app.database import get_sync_db
        from app.services.batch_wordstat_service import check_wordstat_oauth_token
        import asyncio
        loop = asyncio.get_event_loop()
        try:
            token = await loop.run_in_executor(
                None,
                lambda: _check_oauth_token_sync(registry["needs_oauth"]),
            )
            oauth_warning = token is None
        except Exception:
            oauth_warning = True

    return templates.TemplateResponse(
        request,
        f"tools/{slug}/index.html",
        {
            "tool": registry,
            "jobs": jobs,
            "slug": slug,
            "oauth_warning": oauth_warning,
        },
    )


# ---------------------------------------------------------------------------
# 3. POST /{slug}/ — tool submit: validate, create job, dispatch, redirect
# ---------------------------------------------------------------------------
@router.post("/{slug}/", response_class=HTMLResponse, name="tool_submit")
@limiter.limit("10/minute")
async def tool_submit(
    slug: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> RedirectResponse:
    if slug not in TOOL_REGISTRY:
        raise HTTPException(status_code=404, detail="Unknown tool")
    registry = TOOL_REGISTRY[slug]

    form_data = await request.form()
    raw_text = form_data.get(registry["form_field"], "") or ""
    lines = [ln.strip() for ln in raw_text.splitlines() if ln.strip()]

    limit = registry["limit"]
    if len(lines) == 0:
        raise HTTPException(status_code=422, detail="Список не может быть пустым")
    if len(lines) > limit:
        raise HTTPException(status_code=422, detail=f"Превышен лимит: {limit} строк")

    JobModel, _ResultModel = _get_tool_models(slug)

    job_kwargs: dict = {
        "id": uuid.uuid4(),
        "status": "pending",
        "user_id": user.id,
        "created_at": datetime.now(timezone.utc),
        registry["input_col"]: lines,
        registry["count_col"]: len(lines),
    }

    if registry.get("has_domain_field"):
        domain = form_data.get("domain", "") or ""
        job_kwargs["target_domain"] = domain.strip()

    if registry.get("has_region_field"):
        region = form_data.get("region", "213") or "213"
        try:
            job_kwargs["input_region"] = int(region)
        except (ValueError, TypeError):
            job_kwargs["input_region"] = 213

    job = JobModel(**job_kwargs)
    db.add(job)
    await db.commit()
    await db.refresh(job)

    job_id_str = str(job.id)

    if slug == "brief":
        # Brief tool uses a 4-step Celery chain — dispatch via .si() (immutable signatures)
        from app.tasks.brief_tasks import (
            run_brief_step1_serp,
            run_brief_step2_crawl,
            run_brief_step3_aggregate,
            run_brief_step4_finalize,
        )
        celery_chain(
            run_brief_step1_serp.si(job_id_str),
            run_brief_step2_crawl.si(job_id_str),
            run_brief_step3_aggregate.si(job_id_str),
            run_brief_step4_finalize.si(job_id_str),
        ).delay()
    else:
        task_fn = _get_tool_task(slug)
        task_fn.delay(job_id_str)

    return RedirectResponse(f"/ui/tools/{slug}/{job.id}", status_code=303)


# ---------------------------------------------------------------------------
# 4. GET /{slug}/{job_id}/status — HTMX polling partial
# ---------------------------------------------------------------------------
@router.get("/{slug}/{job_id}/status", response_class=HTMLResponse, name="tool_job_status")
async def tool_job_status(
    slug: str,
    job_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> HTMLResponse:
    if slug not in TOOL_REGISTRY:
        raise HTTPException(status_code=404, detail="Unknown tool")
    JobModel, _ResultModel = _get_tool_models(slug)

    stmt = select(JobModel).where(JobModel.id == job_id, JobModel.user_id == user.id)
    result = await db.execute(stmt)
    job = result.scalar_one_or_none()
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")

    return templates.TemplateResponse(
        request,
        f"tools/{slug}/partials/job_status.html",
        {"job": job, "slug": slug},
    )


# ---------------------------------------------------------------------------
# 5. GET /{slug}/{job_id}/export — CSV/XLSX download
# ---------------------------------------------------------------------------
@router.get("/{slug}/{job_id}/export", name="tool_export")
async def tool_export(
    slug: str,
    job_id: uuid.UUID,
    request: Request,
    format: str = Query(default="csv"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> StreamingResponse:
    if slug not in TOOL_REGISTRY:
        raise HTTPException(status_code=404, detail="Unknown tool")
    JobModel, ResultModel = _get_tool_models(slug)

    stmt = select(JobModel).where(JobModel.id == job_id, JobModel.user_id == user.id)
    result = await db.execute(stmt)
    job = result.scalar_one_or_none()
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")

    stmt_results = select(ResultModel).where(ResultModel.job_id == job_id)
    res = await db.execute(stmt_results)
    rows = res.scalars().all()

    headers_row = _EXPORT_HEADERS.get(slug, [])

    # Brief uses a special multi-sheet XLSX export
    if slug == "brief":
        if format != "xlsx":
            raise HTTPException(status_code=400, detail="Brief экспортируется только в формате XLSX")
        brief_result = rows[0] if rows else None
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        # Sheet 1: Title / H1 suggestions
        ws_title = wb.active
        ws_title.title = "Title-H1"
        ws_title.append(["Вариант title / H1"])
        for title in (brief_result.title_suggestions or [] if brief_result else []):
            ws_title.append([title])
        # Sheet 2: H2 cloud
        ws_h2 = wb.create_sheet("H2")
        ws_h2.append(["Заголовок", "Частота"])
        for item in (brief_result.h2_cloud or [] if brief_result else []):
            ws_h2.append([item.get("text", ""), item.get("count", 0)])
        # Sheet 3: Highlights
        ws_hl = wb.create_sheet("Подсветки")
        ws_hl.append(["Подсветка Яндекса"])
        for hl in (brief_result.highlights or [] if brief_result else []):
            ws_hl.append([hl])
        # Sheet 4: Thematic words
        ws_tw = wb.create_sheet("Тематические слова")
        ws_tw.append(["Слово", "Частота"])
        for item in (brief_result.thematic_words or [] if brief_result else []):
            ws_tw.append([item.get("word", ""), item.get("freq", 0)])
        # Sheet 5: Volume stats
        ws_vol = wb.create_sheet("Объём")
        ws_vol.append(["Показатель", "Значение"])
        if brief_result:
            ws_vol.append(["Средняя длина текста (симв.)", brief_result.avg_text_length or 0])
            ws_vol.append(["Среднее число H2", float(brief_result.avg_h2_count or 0)])
            ws_vol.append(["Коммерциализация (%)", brief_result.commercialization_pct or 0])
            ws_vol.append(["Страниц проанализировано", brief_result.pages_crawled or 0])
            ws_vol.append(["Страниц попыток", brief_result.pages_attempted or 0])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"brief_{job_id}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # Wordstat-batch uses a multi-sheet XLSX: Sheet 1 = frequencies, Sheet 2 = monthly dynamics
    if slug == "wordstat-batch":
        if format not in ("xlsx",):
            raise HTTPException(status_code=400, detail="Wordstat-batch экспортируется только в формате XLSX")
        from app.models.wordstat_batch_job import WordstatMonthlyData
        from sqlalchemy import select as sa_select
        wb = openpyxl.Workbook()
        # Sheet 1: Частотность
        ws_freq = wb.active
        ws_freq.title = "Частотность"
        ws_freq.append(["Фраза", "Точная частота", "Широкая частота"])
        for row in rows:
            ws_freq.append([
                getattr(row, "phrase", ""),
                getattr(row, "freq_exact", ""),
                getattr(row, "freq_broad", ""),
            ])
        # Sheet 2: Динамика — join WordstatBatchResult with WordstatMonthlyData
        ws_dyn = wb.create_sheet("Динамика")
        ws_dyn.append(["Фраза", "Месяц", "Частота"])
        result_ids = [row.id for row in rows]
        if result_ids:
            stmt_monthly = sa_select(WordstatMonthlyData).where(
                WordstatMonthlyData.result_id.in_(result_ids)
            )
            monthly_res = await db.execute(stmt_monthly)
            monthly_rows = monthly_res.scalars().all()
            # Build map result_id -> phrase
            id_to_phrase = {row.id: getattr(row, "phrase", "") for row in rows}
            for m in monthly_rows:
                ws_dyn.append([
                    id_to_phrase.get(m.result_id, ""),
                    m.year_month,
                    m.frequency,
                ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"wordstat-batch-{job_id}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if format == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = slug
        ws.append(headers_row)
        for row in rows:
            ws.append(_result_to_row(row, slug))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"{slug}-results-{job_id}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # Default: CSV with UTF-8 BOM
    output = io.StringIO()
    writer = csv.writer(output)
    output.write("\ufeff")  # UTF-8 BOM for Excel compatibility
    writer.writerow(headers_row)
    for row in rows:
        writer.writerow(_result_to_row(row, slug))
    output.seek(0)
    filename = f"{slug}-results-{job_id}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# 6. DELETE /{slug}/{job_id} — delete job + results
# ---------------------------------------------------------------------------
@router.delete("/{slug}/{job_id}", name="tool_delete")
async def tool_delete(
    slug: str,
    job_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    if slug not in TOOL_REGISTRY:
        raise HTTPException(status_code=404, detail="Unknown tool")
    JobModel, ResultModel = _get_tool_models(slug)

    # Verify ownership
    stmt = select(JobModel).where(JobModel.id == job_id, JobModel.user_id == user.id)
    result = await db.execute(stmt)
    job = result.scalar_one_or_none()
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")

    # Delete results first (cascade)
    await db.execute(delete(ResultModel).where(ResultModel.job_id == job_id))
    await db.execute(delete(JobModel).where(JobModel.id == job_id))
    await db.commit()

    headers = {}
    if request.headers.get("HX-Request"):
        headers["HX-Trigger"] = f"jobDeleted-{job_id}"
    return Response(status_code=200, headers=headers)


# ---------------------------------------------------------------------------
# 7. GET /{slug}/{job_id} — results page (LAST — generic catch-all)
# ---------------------------------------------------------------------------
@router.get("/{slug}/{job_id}", response_class=HTMLResponse, name="tool_results")
async def tool_results(
    slug: str,
    job_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> HTMLResponse:
    if slug not in TOOL_REGISTRY:
        raise HTTPException(status_code=404, detail="Unknown tool")
    registry = TOOL_REGISTRY[slug]
    JobModel, ResultModel = _get_tool_models(slug)

    stmt = select(JobModel).where(JobModel.id == job_id, JobModel.user_id == user.id)
    result = await db.execute(stmt)
    job = result.scalar_one_or_none()
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")

    results_rows = []
    brief_result = None
    wordstat_monthly_map: dict[int, list] = {}
    if job.status in ("complete", "partial"):
        stmt_results = select(ResultModel).where(ResultModel.job_id == job_id)
        res = await db.execute(stmt_results)
        results_rows = res.scalars().all()
        if slug == "brief" and results_rows:
            brief_result = results_rows[0]
        if slug == "wordstat-batch" and results_rows:
            from app.models.wordstat_batch_job import WordstatMonthlyData
            result_ids = [r.id for r in results_rows]
            stmt_monthly = select(WordstatMonthlyData).where(
                WordstatMonthlyData.result_id.in_(result_ids)
            )
            monthly_res = await db.execute(stmt_monthly)
            for m in monthly_res.scalars().all():
                wordstat_monthly_map.setdefault(m.result_id, []).append(m)

    return templates.TemplateResponse(
        request,
        f"tools/{slug}/results.html",
        {
            "tool": registry,
            "job": job,
            "results": results_rows,
            "result": brief_result,
            "slug": slug,
            "export_headers": _EXPORT_HEADERS.get(slug, []),
            "wordstat_monthly_map": wordstat_monthly_map,
        },
    )
