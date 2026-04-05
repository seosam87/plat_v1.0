"""Report generation service: dashboard aggregation, PDF, Excel."""
from __future__ import annotations

import io
from datetime import date
from typing import Optional
from uuid import UUID

import openpyxl
from loguru import logger
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ad_traffic import AdTraffic
from app.models.keyword import Keyword
from app.models.project import Project
from app.models.task import SeoTask, TaskStatus


async def dashboard_summary(db: AsyncSession) -> dict:
    """Aggregate dashboard data across all projects."""
    projects = (await db.execute(select(func.count()).select_from(Project))).scalar_one()
    total_keywords = (await db.execute(select(func.count()).select_from(Keyword))).scalar_one()
    open_tasks = (await db.execute(
        select(func.count()).select_from(SeoTask).where(SeoTask.status == TaskStatus.open)
    )).scalar_one()
    in_progress_tasks = (await db.execute(
        select(func.count()).select_from(SeoTask).where(SeoTask.status == TaskStatus.in_progress)
    )).scalar_one()

    return {
        "projects": projects,
        "total_keywords": total_keywords,
        "open_tasks": open_tasks,
        "in_progress_tasks": in_progress_tasks,
    }


async def site_overview(db: AsyncSession, site_id) -> dict:
    """Aggregated overview for a single site: positions, keywords, tasks, crawls."""
    import uuid
    from app.models.crawl import CrawlJob
    from app.models.site import Site

    sid = uuid.UUID(str(site_id))

    keyword_count = (await db.execute(
        select(func.count()).select_from(Keyword).where(Keyword.site_id == sid)
    )).scalar_one()

    open_tasks = (await db.execute(
        select(func.count()).select_from(SeoTask).where(
            SeoTask.site_id == sid, SeoTask.status == TaskStatus.open
        )
    )).scalar_one()

    crawl_count = (await db.execute(
        select(func.count()).select_from(CrawlJob).where(CrawlJob.site_id == sid)
    )).scalar_one()

    # Position distribution
    dist_result = await db.execute(text("""
        WITH latest AS (
            SELECT DISTINCT ON (kp.keyword_id, kp.engine)
                kp.position
            FROM keyword_positions kp
            WHERE kp.site_id = :sid
            ORDER BY kp.keyword_id, kp.engine, kp.checked_at DESC
        )
        SELECT
            COUNT(*) FILTER (WHERE position IS NOT NULL AND position <= 3) AS top3,
            COUNT(*) FILTER (WHERE position IS NOT NULL AND position <= 10) AS top10,
            COUNT(*) FILTER (WHERE position IS NOT NULL AND position <= 30) AS top30,
            COUNT(*) FILTER (WHERE position IS NOT NULL AND position <= 100) AS top100,
            COUNT(*) FILTER (WHERE position IS NULL) AS not_ranked,
            COUNT(*) AS total
        FROM latest
    """), {"sid": sid})
    dist_row = dist_result.mappings().one_or_none()
    distribution = dict(dist_row) if dist_row else {"top3": 0, "top10": 0, "top30": 0, "top100": 0, "not_ranked": 0, "total": 0}

    # Top movers (biggest delta in last 7 days)
    movers_result = await db.execute(text("""
        SELECT DISTINCT ON (kp.keyword_id)
            kp.keyword_id, k.phrase, kp.position, kp.delta, kp.engine
        FROM keyword_positions kp
        JOIN keywords k ON k.id = kp.keyword_id
        WHERE kp.site_id = :sid
          AND kp.delta IS NOT NULL
          AND kp.checked_at >= NOW() - INTERVAL '7 days'
        ORDER BY kp.keyword_id, kp.checked_at DESC
    """), {"sid": sid})
    all_movers = [dict(r) for r in movers_result.mappings().all()]

    gainers = sorted([m for m in all_movers if m["delta"] and m["delta"] > 0], key=lambda x: -x["delta"])[:5]
    losers = sorted([m for m in all_movers if m["delta"] and m["delta"] < 0], key=lambda x: x["delta"])[:5]

    return {
        "keyword_count": keyword_count,
        "open_tasks": open_tasks,
        "crawl_count": crawl_count,
        "distribution": distribution,
        "top_gainers": gainers,
        "top_losers": losers,
    }


def generate_excel_report(
    project_name: str,
    keywords: list[dict],
    tasks: list[dict],
    positions: list[dict],
) -> bytes:
    """Generate a multi-sheet Excel report."""
    wb = openpyxl.Workbook()

    # Positions sheet
    ws_pos = wb.active
    ws_pos.title = "Positions"
    ws_pos.append(["Keyword", "Position", "Delta", "URL", "Engine"])
    for p in positions:
        ws_pos.append([p.get("query", ""), p.get("position"), p.get("delta"), p.get("url", ""), p.get("engine", "")])

    # Keywords sheet
    ws_kw = wb.create_sheet("Keywords")
    ws_kw.append(["Phrase", "Frequency", "Region", "Engine", "Target URL"])
    for k in keywords:
        ws_kw.append([k.get("phrase", ""), k.get("frequency"), k.get("region", ""), k.get("engine", ""), k.get("target_url", "")])

    # Tasks sheet
    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks.append(["Title", "Type", "Status", "URL"])
    for t in tasks:
        ws_tasks.append([t.get("title", ""), t.get("task_type", ""), t.get("status", ""), t.get("url", "")])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


async def ad_traffic_comparison(
    db: AsyncSession,
    site_id,
    period_a_start: date,
    period_a_end: date,
    period_b_start: date,
    period_b_end: date,
) -> list[dict]:
    """Compare ad traffic between two periods."""
    async def _agg(start: date, end: date) -> dict:
        result = await db.execute(
            select(
                AdTraffic.source,
                func.sum(AdTraffic.sessions).label("sessions"),
                func.sum(AdTraffic.conversions).label("conversions"),
                func.sum(AdTraffic.cost).label("cost"),
            )
            .where(AdTraffic.site_id == site_id, AdTraffic.traffic_date.between(start, end))
            .group_by(AdTraffic.source)
        )
        return {r.source: {"sessions": r.sessions, "conversions": r.conversions, "cost": float(r.cost)} for r in result}

    a = await _agg(period_a_start, period_a_end)
    b = await _agg(period_b_start, period_b_end)

    sources = sorted(set(list(a.keys()) + list(b.keys())))
    comparison = []

    def _delta(old, new):
        if old is None or old == 0:
            return None
        return round((new - old) / old * 100, 1)

    for src in sources:
        va = a.get(src, {"sessions": 0, "conversions": 0, "cost": 0.0})
        vb = b.get(src, {"sessions": 0, "conversions": 0, "cost": 0.0})

        cr_a = round(va["conversions"] / va["sessions"] * 100, 2) if va["sessions"] > 0 else 0.0
        cr_b = round(vb["conversions"] / vb["sessions"] * 100, 2) if vb["sessions"] > 0 else 0.0
        cpc_a = round(va["cost"] / va["conversions"], 2) if va["conversions"] > 0 else None
        cpc_b = round(vb["cost"] / vb["conversions"], 2) if vb["conversions"] > 0 else None
        delta_cr_pct = _delta(cr_a, cr_b)
        delta_cpc_pct = _delta(cpc_a, cpc_b) if cpc_a is not None and cpc_b is not None else None

        comparison.append({
            "source": src,
            "period_a": va,
            "period_b": vb,
            "delta_sessions_pct": _delta(va["sessions"], vb["sessions"]),
            "delta_conversions_pct": _delta(va["conversions"], vb["conversions"]),
            "delta_cost_pct": _delta(va["cost"], vb["cost"]),
            "cr_a": cr_a,
            "cr_b": cr_b,
            "delta_cr_pct": delta_cr_pct,
            "cpc_a": cpc_a,
            "cpc_b": cpc_b,
            "delta_cpc_pct": delta_cpc_pct,
        })

    return comparison


async def ad_traffic_trend(
    db: AsyncSession,
    site_id: UUID,
    granularity: str = "weekly",
) -> dict:
    """Return Chart.js-compatible trend data for ad traffic per source.

    Args:
        db: Async database session.
        site_id: Site UUID to filter by.
        granularity: 'weekly' or 'monthly'.

    Returns:
        dict with 'labels' (sorted unique periods as strings) and
        'datasets' (list of {label, data} per source).
    """
    gran = "week" if granularity == "weekly" else "month"

    result = await db.execute(
        text("""
            SELECT source,
                   date_trunc(:gran, traffic_date)::date AS period,
                   SUM(sessions) AS sessions
            FROM ad_traffic
            WHERE site_id = :sid
            GROUP BY source, period
            ORDER BY period
        """),
        {"gran": gran, "sid": str(site_id)},
    )
    rows = result.mappings().all()

    # Collect all unique periods and sources
    periods_set: set[str] = set()
    sources_set: set[str] = set()
    data_map: dict[str, dict[str, int]] = {}  # source -> {period_str -> sessions}

    for row in rows:
        period_str = str(row["period"])
        src = row["source"]
        sessions = int(row["sessions"] or 0)
        periods_set.add(period_str)
        sources_set.add(src)
        data_map.setdefault(src, {})[period_str] = sessions

    labels = sorted(periods_set)
    sources = sorted(sources_set)

    # Tailwind-palette hex colors for up to 4 sources
    colors = ["#6366f1", "#10b981", "#f59e0b", "#ef4444"]

    datasets = []
    for idx, src in enumerate(sources):
        color = colors[idx % len(colors)]
        src_data = data_map.get(src, {})
        datasets.append({
            "label": src,
            "data": [src_data.get(p, 0) for p in labels],
            "borderColor": color,
            "backgroundColor": color + "33",  # 20% opacity fill
            "tension": 0.3,
        })

    return {"labels": labels, "datasets": datasets}
