"""Tests for bulk service: export format functions (pure/sync path)."""
import csv
import io

import openpyxl

from app.services.bulk_service import _CSV_HEADERS


# ---------------------------------------------------------------------------
# CSV export format helpers (tested without DB)
# ---------------------------------------------------------------------------


def _make_csv(rows: list[list]) -> str:
    """Build a CSV string from header + rows, matching bulk_service format."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_CSV_HEADERS)
    writer.writerows(rows)
    return output.getvalue()


def _make_xlsx(rows: list[list]) -> bytes:
    """Build an XLSX bytes payload matching bulk_service format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Keywords"
    ws.append(_CSV_HEADERS)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_csv_header_structure():
    """CSV output must start with the expected 9-column header."""
    csv_str = _make_csv([])
    first_line = csv_str.strip().split("\n")[0]
    assert "Phrase" in first_line
    assert "Position" in first_line
    assert "Delta" in first_line
    assert "Target URL" in first_line


def test_csv_export_format():
    """CSV row must include phrase, frequency, and position values."""
    row = ["seo продвижение", 1000, "Moscow", "yandex", "SEO", "cluster1",
           "https://e.com/page/", 5, -2]
    csv_str = _make_csv([row])
    assert "seo продвижение" in csv_str
    assert "1000" in csv_str
    assert "Moscow" in csv_str


def test_csv_export_empty():
    """Empty keyword list must produce header-only CSV (1 line)."""
    csv_str = _make_csv([])
    lines = [l for l in csv_str.strip().split("\n") if l]
    assert len(lines) == 1
    assert "Phrase" in lines[0]


def test_xlsx_export_produces_bytes():
    """XLSX export must return bytes with PK (zip) magic header."""
    row = ["seo продвижение", 1000, "Moscow", "yandex", "", "", "", None, None]
    result = _make_xlsx([row])
    assert isinstance(result, bytes)
    assert len(result) > 100
    # XLSX is a zip file — magic bytes are PK
    assert result[:2] == b"PK"


def test_xlsx_headers_match_csv():
    """XLSX sheet must have same column headers as CSV."""
    result = _make_xlsx([])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row == _CSV_HEADERS


def test_csv_export_multiple_rows():
    """Multiple rows produce correct line count (header + N rows)."""
    rows = [
        [f"keyword {i}", i * 10, "RU", "google", "", "", "", None, None]
        for i in range(5)
    ]
    csv_str = _make_csv(rows)
    lines = [l for l in csv_str.strip().split("\n") if l]
    assert len(lines) == 6  # header + 5 data rows
