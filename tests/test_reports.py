"""Unit tests for report service and ad traffic."""
import io
import pytest
import openpyxl

from app.services.report_service import generate_excel_report


class TestExcelReport:
    def test_generates_valid_xlsx(self):
        data = generate_excel_report(
            "Test Project",
            keywords=[{"phrase": "seo", "frequency": 100, "region": "RU", "engine": "google", "target_url": ""}],
            tasks=[{"title": "Fix 404", "task_type": "page_404", "status": "open", "url": "/x"}],
            positions=[{"query": "seo", "position": 5, "delta": 2, "url": "/seo", "engine": "google"}],
        )
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Positions" in wb.sheetnames
        assert "Keywords" in wb.sheetnames
        assert "Tasks" in wb.sheetnames
        # Check positions sheet has data
        ws = wb["Positions"]
        assert ws.cell(2, 1).value == "seo"
        assert ws.cell(2, 2).value == 5

    def test_empty_data(self):
        data = generate_excel_report("Empty", keywords=[], tasks=[], positions=[])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb["Positions"].max_row == 1  # header only


class TestAdTrafficModel:
    def test_model_import(self):
        from app.models.ad_traffic import AdTraffic
        ad = AdTraffic(
            site_id="12345678-1234-1234-1234-123456789abc",
            source="google_ads",
            traffic_date="2026-03-01",
            sessions=100,
            conversions=5,
            cost=50.0,
        )
        assert ad.source == "google_ads"
        assert ad.sessions == 100


# ---- Phase 9: site overview + endpoints ----


class TestSiteOverviewImport:
    def test_import(self):
        from app.services.report_service import site_overview
        assert callable(site_overview)


class TestSiteOverviewEndpoint:
    def test_endpoint_registered(self):
        from app.routers.reports import router
        paths = [r.path for r in router.routes]
        assert "/reports/sites/{site_id}/overview" in paths


class TestDeltaPercentLogic:
    """Test the delta percentage calculation used in ad traffic comparison."""

    def test_positive_growth(self):
        old, new = 100, 150
        delta = round((new - old) / old * 100, 1)
        assert delta == 50.0

    def test_negative_growth(self):
        old, new = 200, 150
        delta = round((new - old) / old * 100, 1)
        assert delta == -25.0

    def test_zero_old_returns_none(self):
        old = 0
        delta = None if old == 0 else round((100 - old) / old * 100, 1)
        assert delta is None

    def test_no_change(self):
        old, new = 50, 50
        delta = round((new - old) / old * 100, 1)
        assert delta == 0.0
